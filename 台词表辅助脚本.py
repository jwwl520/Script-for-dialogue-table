import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import win32com.client as win32 
from datetime import datetime
import pythoncom
from openpyxl import load_workbook 
import difflib
import traceback
import sys

# --- COM Constants ---
MSO_TRUE = -1 
MSO_FALSE = 0 
MSO_SCALE_FROM_TOP_LEFT = 0

# Shape types
MSO_PICTURE = 13
MSO_LINKED_PICTURE = 16
MSO_PLACEHOLDER = 14

class ExcelBatchProcessor:
    def __init__(self, root):
        self.root = root
        self.running = False
        self.setup_config()
        self.setup_ui()
        self.excel = None
        self.file_queue = []
        self.old_queue = []

    def setup_config(self):
        self.config = {
            'col_e': 'E', 'col_g': 'G', 'col_f_speaker': 'F',
            'chinese_filter': True, 'merge_duplicates': True,
            'copy_intro': True, 'copy_speakers': True,
            'speaker_match_threshold': 0.6, 'exact_match_threshold': 0.95,
            'exact_match_length_ratio_threshold': 0.7,
            'segment_coherence_weight': 0.25, 'segment_length_ratio_weight': 0.15,
            'segment_similarity_weight': 0.60, 
            'content_guess_confidence_threshold': 0.7,
            'old_file_match_threshold': 0.7,
            'image_scale_factor': 0.9,
            'last_folder': os.getcwd(), 'output_folder': os.getcwd(),
            'log_level': "INFO"
        }

    def setup_ui(self):
        self.root.title("Excel高级批量处理器 v8.16 (UI属性修复)") 
        self.root.geometry("900x750")

        top_controls_frame = ttk.Frame(self.root)
        top_controls_frame.pack(pady=10, padx=10, fill=tk.X)

        ttk.Label(top_controls_frame, text="新表:").grid(row=0, column=0, padx=(0,5), pady=3, sticky='w')
        ttk.Button(top_controls_frame, text="选择新表文件", command=self.select_file).grid(row=0, column=1, padx=2, pady=3, sticky='ew')
        ttk.Button(top_controls_frame, text="选择新表文件夹", command=self.select_folder).grid(row=0, column=2, padx=2, pady=3, sticky='ew')

        ttk.Label(top_controls_frame, text="旧表:").grid(row=1, column=0, padx=(0,5), pady=3, sticky='w')
        ttk.Button(top_controls_frame, text="选择旧表文件", command=self.select_old_file).grid(row=1, column=1, padx=2, pady=3, sticky='ew')
        ttk.Button(top_controls_frame, text="选择旧表文件夹", command=self.select_old_folder).grid(row=1, column=2, padx=2, pady=3, sticky='ew')

        ttk.Button(top_controls_frame, text="选择输出位置", command=self.select_output_folder).grid(row=2, column=0, columnspan=1, padx=(0,2), pady=(10,3), sticky='ew')

        s = ttk.Style(); s.configure("Accent.TButton", font=('Segoe UI',9,'bold'), foreground='green'); s.configure("Stop.TButton", font=('Segoe UI',9,'bold'), foreground='red')
        ttk.Button(top_controls_frame, text="开始处理", command=self.start_processing, style="Accent.TButton").grid(row=2, column=1, padx=2, pady=(10,3), ipady=4, sticky='ew')
        ttk.Button(top_controls_frame, text="停止", command=self.stop_processing, style="Stop.TButton").grid(row=2, column=2, padx=2, pady=(10,3), ipady=4, sticky='ew')

        top_controls_frame.columnconfigure(1, weight=1); top_controls_frame.columnconfigure(2, weight=1)

        cfg_frame = ttk.LabelFrame(self.root, text="处理配置")
        cfg_frame.pack(pady=5, padx=10, fill=tk.X)

        # E, G, F 列输入框 - 使用 setattr 创建 self.entry_e, self.entry_g, self.entry_speaker
        cols_cfg_data = [("内容列(E):", 'col_e'), ("文本列(G):", 'col_g'), ("说话人列(F):", 'col_f_speaker')]
        for i, (text, cfg_key) in enumerate(cols_cfg_data):
            ttk.Label(cfg_frame, text=text).grid(row=0, column=i*2, padx=5, pady=3, sticky='w')
            entry = ttk.Entry(cfg_frame, width=5)
            entry.insert(0, self.config[cfg_key])
            entry.grid(row=0, column=i*2+1, padx=5, pady=3)
            # setattr f"entry_{cfg_key.split('_')[-1]}" 会创建 self.entry_e, self.entry_g, self.entry_speaker
            setattr(self, f"entry_{cfg_key.split('_')[-1]}", entry) 

        # 复选框配置 - 使用明确的属性名 self.var_ch, self.var_mg, self.var_int, self.var_spk
        self.var_ch = tk.BooleanVar(value=self.config['chinese_filter'])
        ttk.Checkbutton(cfg_frame, text="启用中文过滤", variable=self.var_ch).grid(row=1, column=0, columnspan=2, padx=5, pady=3, sticky='w')
        
        self.var_mg = tk.BooleanVar(value=self.config['merge_duplicates'])
        ttk.Checkbutton(cfg_frame, text="启用合并重复", variable=self.var_mg).grid(row=1, column=2, columnspan=2, padx=5, pady=3, sticky='w')
        
        self.var_int = tk.BooleanVar(value=self.config['copy_intro'])
        ttk.Checkbutton(cfg_frame, text="启用简介复制", variable=self.var_int).grid(row=1, column=4, columnspan=2, padx=5, pady=3, sticky='w')
        
        self.var_spk = tk.BooleanVar(value=self.config['copy_speakers'])
        ttk.Checkbutton(cfg_frame, text="启用说话人复制", variable=self.var_spk).grid(row=2, column=0, columnspan=2, padx=5, pady=3, sticky='w')

        # 阈值输入框 - 使用明确的属性名 self.entry_th_speaker, self.entry_th_filename
        ttk.Label(cfg_frame, text="内容匹配阈值:").grid(row=2, column=2, padx=5, pady=3, sticky='w')
        self.entry_th_speaker = ttk.Entry(cfg_frame, width=5)
        self.entry_th_speaker.insert(0, str(self.config['speaker_match_threshold']))
        self.entry_th_speaker.grid(row=2, column=3, padx=5, pady=3)
        
        ttk.Label(cfg_frame, text="文件名匹配阈值:").grid(row=2, column=4, padx=5, pady=3, sticky='w')
        self.entry_th_filename = ttk.Entry(cfg_frame, width=5)
        self.entry_th_filename.insert(0, str(self.config['old_file_match_threshold']))
        self.entry_th_filename.grid(row=2, column=5, padx=5, pady=3)

        log_frame = ttk.LabelFrame(self.root, text="日志输出")
        log_frame.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def log(self, message, level="INFO"):
        level_map={"TRACE":0,"DEBUG":1,"INFO":2,"WARNING":3,"ERROR":4,"CRITICAL":5}
        cfg_log_lvl_str=self.config.get('log_level',"INFO").upper()
        cfg_log_lvl=level_map.get(cfg_log_lvl_str,2)
        msg_lvl=level_map.get(level.upper(),2)
        if msg_lvl>=cfg_log_lvl:
            ts=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.log_text.insert(tk.END,f"[{level.upper()}] {ts} - {message}\n")
            self.log_text.see(tk.END)
            if self.root and self.root.winfo_exists(): self.root.update_idletasks()

    def log_with_context(self, message, row=None, col_idx=None, level="INFO"):
        ctx_info=""
        if row is not None: ctx_info=f"行{row}"
        if col_idx is not None:
            try:col_ltr=self._idx2col(int(col_idx))
            except:col_ltr=str(col_idx)
            ctx_info+=f":{col_ltr}"
        full_msg=f"{ctx_info} - {message}" if ctx_info else message
        self.log(full_msg,level)

    def select_file(self):
        if self.running: messagebox.showwarning("提示","处理中..."); return
        fp=filedialog.askopenfilename(title="选新表",filetypes=[("Excel","*.xlsx")],initialdir=self.config.get('last_folder',os.getcwd()))
        if fp:self.config['last_folder']=os.path.dirname(fp);self.file_queue=[fp];self.log(f"选新表(单):{os.path.basename(fp)}")

    def select_folder(self):
        if self.running: messagebox.showwarning("提示","处理中..."); return
        dp=filedialog.askdirectory(title="选新表文件夹",initialdir=self.config.get('last_folder',os.getcwd()))
        if dp:
            self.config['last_folder']=dp
            self.file_queue=[os.path.join(dp,f) for f in os.listdir(dp) if f.lower().endswith('.xlsx') and not f.startswith('~')]
            self.log(f"选新表文件夹:{dp}({len(self.file_queue)}个)" if self.file_queue else f"新表文件夹{dp}无.xlsx","INFO" if self.file_queue else "WARNING")

    def select_old_file(self):
        if self.running: messagebox.showwarning("提示","处理中..."); return
        fp=filedialog.askopenfilename(title="选旧表",filetypes=[("Excel","*.xlsx")],initialdir=self.config.get('last_folder',os.getcwd()))
        if fp:self.old_queue=[fp];self.log(f"选旧表(单):{os.path.basename(fp)}")

    def select_old_folder(self):
        if self.running: messagebox.showwarning("提示","处理中..."); return
        dp=filedialog.askdirectory(title="选旧表文件夹",initialdir=self.config.get('last_folder',os.getcwd()))
        if dp:
            self.old_queue=[os.path.join(dp,f) for f in os.listdir(dp) if f.lower().endswith('.xlsx') and not f.startswith('~')]
            self.log(f"选旧表文件夹:{dp}({len(self.old_queue)}个)" if self.old_queue else f"旧表文件夹{dp}无.xlsx","INFO" if self.old_queue else "WARNING")

    def select_output_folder(self):
        dp=filedialog.askdirectory(title="选输出位置",initialdir=self.config.get('output_folder',os.getcwd()))
        if dp:self.config['output_folder']=dp;self.log(f"输出设为:{dp}")

    def start_processing(self):
        if self.running:messagebox.showwarning("处理中...","当前任务进行中。");return
        if not self.file_queue:messagebox.showerror("错误","请先选新表。");return
        self.update_cfg_from_ui() # 调用修正后的配置读取方法
        if self.config.get('copy_speakers') and not self.old_queue:messagebox.showerror("错误","启用说话人复制但未选旧表。");return
        self.running=True;self.log("===== 处理开始 =====","INFO")
        try:
            pythoncom.CoInitialize()
            self.excel=win32.gencache.EnsureDispatch('Excel.Application')
            self.excel.Visible=False;self.excel.DisplayAlerts=False
            for i,fp in enumerate(self.file_queue):
                if not self.running:self.log("处理被中止。","INFO");break
                self.log(f"--- 处理第{i+1}/{len(self.file_queue)}个文件:{os.path.basename(fp)} ---","INFO")
                self._proc_file(fp)
            if self.running:self.log("所有文件处理完毕。","INFO");messagebox.showinfo("完成","所有文件处理完毕！")
        except Exception as e:
            self.log(f"主流程严重错误:{e}","CRITICAL");self.log(traceback.format_exc(),"DEBUG")
            messagebox.showerror("严重错误",f"发生错误，详见日志:{e}")
        finally:
            if self.excel:
                try:self.excel.Quit()
                except Exception as eq:self.log(f"关Excel出错:{eq}","ERROR")
                self.excel=None
            pythoncom.CoUninitialize()
            self.running=False;self.log("===== 处理结束 =====","INFO")

    def stop_processing(self):
        if self.running:self.running=False;self.log("用户请求停止...");messagebox.showinfo("停止请求","将尝试停止。")
        else:self.log("当前无处理任务。","INFO")

    def update_cfg_from_ui(self):
        try:
            self.config.update({
                'col_e': self.entry_e.get().upper(),             # 使用 self.entry_e
                'col_g': self.entry_g.get().upper(),             # 使用 self.entry_g
                'col_f_speaker': self.entry_speaker.get().upper(),# 使用 self.entry_speaker
                'chinese_filter': self.var_ch.get(),             # 使用 self.var_ch
                'merge_duplicates': self.var_mg.get(),           # 使用 self.var_mg
                'copy_intro': self.var_int.get(),                # 使用 self.var_int
                'copy_speakers': self.var_spk.get(),             # 使用 self.var_spk (这个之前是正确的)
                'speaker_match_threshold': float(self.entry_th_speaker.get()),
                'old_file_match_threshold': float(self.entry_th_filename.get())
            })
            log_cfg={k:v for k,v in self.config.items() if k not in ['last_folder','output_folder']}
            self.log(f"配置已从UI更新: {log_cfg}", "DEBUG")
        except ValueError: # 特别是float转换
            self.log("阈值输入无效，请输入数字。", "ERROR"); messagebox.showerror("配置错误", "匹配阈值必须是数字。")
            # 保留之前的有效值或设置一个安全的默认值
            self.config['speaker_match_threshold'] = self.config.get('speaker_match_threshold', 0.6)
            self.config['old_file_match_threshold'] = self.config.get('old_file_match_threshold', 0.7)
        except AttributeError as ae: 
             self.log(f"UI配置读取错误: {ae}. 请检查UI控件变量名是否与代码中使用的属性名一致。", "CRITICAL")
             messagebox.showerror("UI错误", f"读取配置控件时出错: {ae}")


    def _normalize_name_for_matching(self, filename_no_ext):
        name = re.sub(r'[【《\(（\[].*?[】》\)）\]]', '', filename_no_ext) 
        name = name.replace("【】", "") 
        common_affixes = [" - 译制台词表", "译制台词表", " - subtitles", "subtitles", " - 副本", "副本"]
        for affix in common_affixes:
            if name.lower().endswith(affix.lower()): name = name[:-len(affix)]
            if name.lower().startswith(affix.lower()): name = name[len(affix):]
        name = name.replace("仅字幕", "").replace(" ", "").lower().strip('-_. ')
        name = re.sub(r'[-_]', '', name)
        return name

    def _match_old(self, new_filepath):
        # 防止新旧表相同
        for old_fp in self.old_queue:
            if os.path.normpath(os.path.abspath(old_fp)) == os.path.normpath(os.path.abspath(new_filepath)):
                self.log(f"警告: 新表和旧表是同一个文件 '{os.path.basename(old_fp)}'，无法从自身复制", "WARNING")
                return None
        
        # 原有的匹配逻辑...
        base_new_orig = os.path.basename(new_filepath)
        norm_new = self._normalize_name_for_matching(os.path.splitext(base_new_orig)[0])
        if not norm_new: self.log(f"新文件名 '{base_new_orig}' 规范化后为空。", "WARNING"); return None
        
        best_path, best_sim = None, 0.0
        if not self.old_queue: self.log("旧表队列为空。", "WARNING"); return None
        self.log(f"新表规范名: '{norm_new}' (来自 '{base_new_orig}')", "DEBUG")

        for old_fp in self.old_queue:
            norm_old = self._normalize_name_for_matching(os.path.splitext(os.path.basename(old_fp))[0])
            if not norm_old: self.log(f"旧文件名 '{os.path.basename(old_fp)}' 规范化后为空, 跳过。", "TRACE"); continue
            sim = difflib.SequenceMatcher(None, norm_new, norm_old).ratio()
            self.log(f"  比较旧表: '{norm_old}' (来自 '{os.path.basename(old_fp)}') vs '{norm_new}' -> 相似度: {sim:.3f}", "TRACE")
            if sim > best_sim: best_sim = sim; best_path = old_fp
        
        threshold = self.config.get('old_file_match_threshold', 0.7)
        if best_sim >= threshold:
            self.log(f"为 '{base_new_orig}' 找到旧表: '{os.path.basename(best_path)}' (相似度: {best_sim:.3f})", "INFO")
            return best_path
        else:
            self.log(f"未能为 '{base_new_orig}' 找到足够相似旧表 (最高相似度 {best_sim:.3f} < 阈值 {threshold})", "WARNING")
            return None

    def _proc_file(self, newp):
        name = os.path.basename(newp)
        start_time = datetime.now()
        current_stage = "初始化"
        wb, ws = None, None
        g_col_idx_local = self._col2idx(self.config['col_g']) # 在try块外先定义，确保finally中可用(如果需要)

        def detect_start_row(sheet_obj, col_idx, sheet_name_log):
            # ... (此函数实现与上一版相同)
            self.log(f"检测 '{sheet_name_log}' (列 {self._idx2col(col_idx)}) 数据起始行...", "DEBUG")
            for r_detect in range(1, 16):
                try:
                    val = str(sheet_obj.Cells(r_detect, col_idx).Value or "").lower()
                    if val and any(h in val for h in ["序号", "台词", "说话人", "line id", "dialogue", "start time"]):
                        self.log(f"'{sheet_name_log}': 检测到表头 '{val}' 在行 {r_detect}, 数据从行 {r_detect+1} 开始。", "INFO")
                        return r_detect + 1
                except Exception as e_detect_cell: 
                    self.log(f"检测起始行单元格({r_detect},{col_idx})出错: {e_detect_cell}", "TRACE")
            d_start = 5; self.log(f"'{sheet_name_log}': 未检测到表头 (检查前15行), 默认从行 {d_start} 开始。", "INFO"); return d_start

        try:
            current_stage = "打开文件"; wb = self.excel.Workbooks.Open(newp)
            current_stage = "获取工作表"
            try: ws = wb.ActiveSheet; self.log(f"打开 '{name}', 活动表: '{ws.Name}'", "INFO")
            except: 
                if wb.Sheets.Count > 0: ws = wb.Sheets(1); self.log(f"'{name}': 用首个表 '{ws.Name}'", "WARNING")
                else: self.log(f"'{name}': 无工作表!", "ERROR"); wb.Close(False); return

            # g_col_idx_local 在上面定义了
            DATA_START_ROW = detect_start_row(ws, g_col_idx_local, ws.Name)

            current_stage = "匹配旧表"; old_path = self._match_old(newp)
            # 修正：使用 self.var_spk (这是在 setup_ui 中创建的 BooleanVar 的正确属性名)
            if self.config.get('copy_speakers') and self.var_spk.get() and not old_path: 
                self.log(f"警告: 为 '{name}' 启用说话人复制但未找到旧表。", "WARNING")

            # 修正：使用 self.var_int
            if self.config.get('copy_intro') and self.var_int.get() and old_path:
                current_stage = "复制简介"; self._copy_intro(ws, old_path)
            if not self.running: self.log("中止于简介后"); wb.Close(False); return

            # 修正：使用 self.var_spk
            if self.config.get('copy_speakers') and self.var_spk.get() and old_path: 
                current_stage = "复制说话人"; self._copy_speakers(ws, old_path, DATA_START_ROW)
            if not self.running: self.log("中止于说话人复制后"); wb.Close(False); return

            current_stage = "内容清洗/合并"; last_row = DATA_START_ROW - 1
            try: last_row = ws.Cells.Find("*", SearchOrder=win32.constants.xlByRows, SearchDirection=win32.constants.xlPrevious).Row
            except: self.log(f"'{ws.Name}': 清洗前无法确定最后行", "WARNING")

            r_loop_local = DATA_START_ROW # 初始化循环变量，确保在except块中可引用
            if last_row >= DATA_START_ROW:
                rows_del, merge_data = [], {}
                e_col_idx = self._col2idx(self.config['col_e'])
                for r_loop_local in range(DATA_START_ROW, last_row + 1):
                    if not self.running: break
                    g_val = str(ws.Cells(r_loop_local, g_col_idx_local).Value or "").strip()
                    del_reason, should_del = "", False
                    merge_key_base = ''.join(re.findall(r'[\u4e00-\u9fff。？！]', g_val)) if self.config.get('chinese_filter') else g_val
                    merge_data[r_loop_local] = re.sub(r'([。？！])\1{2,}', r'\1\1', merge_key_base) if self.config.get('chinese_filter') and merge_key_base else merge_key_base
                    
                    if self.config.get('chinese_filter'): # 使用 .get() 避免KeyError
                        if not re.search(r'[\u4e00-\u9fff]', g_val) and len(set(re.findall(r'[A-Za-z]', g_val))) < 3:
                            should_del, del_reason = True, "中文过滤:无中文且字母少于3种"
                    elif len(set(re.findall(r'[A-Za-z]', g_val))) < 3 and not re.search(r'[\u4e00-\u9fff0-9]', g_val):
                        should_del, del_reason = True, "常规:字母少于3种且无中文或数字"
                    if should_del: rows_del.append(r_loop_local); self.log_with_context(f"标记删除({del_reason}):'{g_val[:30]}'", r_loop_local, g_col_idx_local, "DEBUG")
                if not self.running: self.log("中止于清洗判断后"); wb.Close(False); return

                if self.config.get('merge_duplicates'):
                    groups = {}
                    for r_num, key in sorted(merge_data.items()):
                        if r_num in rows_del or not key: continue
                        if key not in groups: groups[key] = []
                        groups[key].append(r_num)
                    for key, g_rows in groups.items():
                        if not self.running: break
                        if len(g_rows) > 1:
                            try:
                                ws.Cells(g_rows[0], e_col_idx).Value = ws.Cells(g_rows[-1], e_col_idx).Value
                                rows_del.extend(g_rows[1:])
                                self.log(f"合并'{key[:20]}..':保留行{g_rows[0]},E列来自行{g_rows[-1]},删{g_rows[1:]}", "DEBUG")
                            except Exception as me: self.log(f"合并'{key[:20]}..'出错:{me}", "WARNING")
                if not self.running: self.log("中止于合并后"); wb.Close(False); return

                if rows_del:
                    unique_del = sorted(list(set(rows_del)), reverse=True)
                    self.log(f"准备删除 {len(unique_del)} 行...", "INFO")
                    for r_d in unique_del:
                        if not self.running: break
                        try: ws.Rows(r_d).Delete()
                        except Exception as de: self.log_with_context(f"删除失败:{de}", r_d, level="WARNING")
                else: self.log("内容清洗:无行标记删除", "INFO")
            else: self.log(f"'{ws.Name}': 数据行({last_row})<起始行({DATA_START_ROW}),跳过清洗", "INFO")
            if not self.running: self.log("中止于删除行后"); wb.Close(False); return
            
            current_stage = "段落标点调整"; final_lr_para_punct = DATA_START_ROW - 1
            try: final_lr_para_punct = ws.Cells.Find("*", SearchOrder=win32.constants.xlByRows, SearchDirection=win32.constants.xlPrevious).Row
            except: self.log(f"'{ws.Name}': 段落标点前无法确定最后行", "WARNING")
            if final_lr_para_punct >= DATA_START_ROW: self._apply_default_punctuation_to_g_column(ws, final_lr_para_punct, DATA_START_ROW)
            else: self.log(f"'{ws.Name}': 段落标点前数据行不足,跳过", "INFO")
            if not self.running: self.log("中止于段落标点后"); wb.Close(False); return

            current_stage = "最终标点检查"; final_lr_final_punct = DATA_START_ROW - 1
            try: final_lr_final_punct = ws.Cells.Find("*", SearchOrder=win32.constants.xlByRows, SearchDirection=win32.constants.xlPrevious).Row
            except: self.log(f"'{ws.Name}': 最终标点前无法确定最后行", "WARNING")
            if final_lr_final_punct >= DATA_START_ROW: self._final_ensure_punctuation(ws, final_lr_final_punct, DATA_START_ROW, g_col_idx_local)
            else: self.log(f"'{ws.Name}': 最终标点前数据行不足,跳过", "INFO")
            if not self.running: self.log("中止于最终标点后"); wb.Close(False); return

            current_stage = "调整图片"; self.adjust_images(ws)
            if not self.running: self.log("中止于图片调整后"); wb.Close(False); return

            current_stage = "保存文件"; out_name = f"改_{name}"
            out_path = os.path.normpath(os.path.join(self.config['output_folder'], out_name))
            try: wb.SaveAs(out_path, FileFormat=51); self.log(f"✅ 保存到: {out_path} [耗时: {(datetime.now()-start_time).total_seconds():.2f}s]")
            except Exception as se:
                self.log(f"保存 '{out_path}' 失败: {se}. 尝试备用名...", "ERROR")
                try:
                    bk_path = os.path.join(self.config['output_folder'], f"改_{os.path.splitext(name)[0]}_{datetime.now().strftime('%H%M%S')}{os.path.splitext(name)[1]}")
                    wb.SaveAs(bk_path, FileFormat=51); self.log(f"✅ 已用备用名保存: {bk_path}")
                except Exception as sbe: self.log(f"备用名保存也失败: {sbe}", "CRITICAL")
        except Exception as epf:
            ec_list = traceback.extract_tb(sys.exc_info()[2]); ef = ec_list[-1] if ec_list else None
            efn,eln,efunc = (os.path.basename(ef.filename),ef.lineno,ef.name) if ef else ("N/A",)*3
            lmsg = f"处理'{name}'阶段'{current_stage}'失败:{epf} ({efn}:{eln} {efunc})"
            self.log(lmsg, "ERROR"); self.log(traceback.format_exc(), "DEBUG")
            if 'r_loop_local' in locals() and isinstance(r_loop_local,int) and ws: 
                try: 
                    g_val_err = str(ws.Cells(r_loop_local,g_col_idx_local).Value or '')[:30]
                    f_val_err = str(ws.Cells(r_loop_local,self._col2idx(self.config['col_f_speaker'])).Value or '')[:20]
                    self.log_with_context(f"错误可能在行附近. G='{g_val_err}', F='{f_val_err}'",r_loop_local,level="DEBUG")
                except Exception as e_ctx_log: self.log(f"记录错误上下文时出错: {e_ctx_log}", "WARNING")
        finally: # _proc_file 的 finally
            if wb: 
                try: wb.Close(SaveChanges=False)
                except Exception as ec: self.log(f"关闭工作簿 '{name}' 出错:{ec}", "ERROR")

    def _get_trailing_punctuation(self, text_input):
        if not text_input or not isinstance(text_input, str): return ""
        m = re.search(r'([。！？，.,!?:;\uff0c\uff1b\uff1a\uff1f\uff01\uff0e\s]+)$', text_input)
        return m.group(1).strip() if m else ""

    def _apply_default_punctuation_to_g_column(self, ws_punct, last_valid_row, data_start_row_punct):
        if not ws_punct or last_valid_row < data_start_row_punct: self.log(f"段落标点: 无效参数或数据不足 (行{data_start_row_punct}-{last_valid_row})", "WARNING"); return
        self.log(f"段落标点调整 (行 {data_start_row_punct}-{last_valid_row})...", "INFO")
        g_col, f_col = self._col2idx(self.config['col_g']), self._col2idx(self.config['col_f_speaker'])
        valid_end_puncts = ('。','！','？','.','!','?',';','；',':','：',')','）','"','\'','”','’',']','】','》','…','>','}','\uff0e','\uff01','\uff1f','\uff1b','\uff1a','\uff09','\uff02','\u201d','\u2019','\u300b','\u2026')
        modified, paras = 0, []; curr_para_rows, curr_speaker = [], "<INIT_SPEAKER_ADP>"
        for r_p in range(data_start_row_punct, last_valid_row + 1):
            if not self.running: break
            try:
                spk = str(ws_punct.Cells(r_p, f_col).Value or "").strip()
                dlg = str(ws_punct.Cells(r_p, g_col).Value or "").strip()
                if not dlg: 
                    if curr_para_rows: paras.append(list(curr_para_rows)); curr_para_rows.clear()
                    curr_speaker = "<RESET_SPEAKER_EMPTY_DLG_ADP>" 
                    continue
                if spk != curr_speaker:
                    if curr_para_rows: paras.append(list(curr_para_rows)); curr_para_rows.clear()
                    curr_speaker = spk
                curr_para_rows.append(r_p)
            except Exception as epd: self.log_with_context(f"段落识别错:{epd}",r_p,level="WARNING")
        if curr_para_rows: paras.append(list(curr_para_rows))
        if not paras: self.log("段落标点:未识别到段落", "INFO"); return
        self.log(f"段落标点:识别到 {len(paras)} 段", "DEBUG")
        for para_r_list in paras:
            if not self.running or not para_r_list: break
            num_lines = len(para_r_list)
            for i_line, r_line_val in enumerate(para_r_list):
                try:
                    cell = ws_punct.Cells(r_line_val, g_col); text = str(cell.Value or "").strip()
                    if not text: continue
                    is_last, has_punct = (i_line == num_lines - 1), any(text.endswith(p) for p in valid_end_puncts)
                    if is_last and not has_punct: cell.Value = text + "。"; modified +=1; self.log_with_context(f"段落标点:末行'{text[:20]}'加。",r_line_val,g_col,"TRACE")
                    elif not is_last and has_punct:
                        orig_t = text
                        while any(text.endswith(p) for p in valid_end_puncts) and text:
                            changed_in_iter = False
                            for p_rem in valid_end_puncts: 
                                if text.endswith(p_rem): text = text[:-len(p_rem)].strip(); changed_in_iter=True; break
                            if not changed_in_iter: break 
                        if text != orig_t: cell.Value = text; modified+=1; self.log_with_context(f"段落标点:中行'{orig_t[:20]}'去标点为'{text[:20]}'",r_line_val,g_col,"TRACE")
                except Exception as epc: self.log_with_context(f"段落标点单元格操作错:{epc}",r_line_val,g_col,"WARNING")
        self.log(f"段落标点调整完毕,修改 {modified} 处", "INFO")

    def _final_ensure_punctuation(self, ws, last_row, data_start_row, g_col_idx):
        if not ws or last_row < data_start_row: self.log("最终标点:无效参数或数据不足", "INFO"); return
        self.log(f"最终标点检查 (行 {data_start_row}-{last_row})...", "INFO")
        modified_count = 0
        valid_endings = ('。','！','？','.','!','?',';','；',':','：',')','）','"','\'','”','’',']','】','》','…','>','}','\uff0e','\uff01','\uff1f','\uff1b','\uff1a','\uff0c','\uff09','\u201d','\u2019','\u300b','\u2026',',')

        for r_fep in range(data_start_row, last_row + 1):
            if not self.running: break
            try:
                cell = ws.Cells(r_fep, g_col_idx)
                text = str(cell.Value or '').strip()
                if text and not any(text.endswith(p) for p in valid_endings):
                    new_val = text + "。"
                    cell.Value = new_val; modified_count += 1
                    self.log_with_context(f"最终标点: 为 '{text[:30]}' 加句号 -> '{new_val[:31]}'", r_fep, g_col_idx, "TRACE")
            except Exception as efp_cell: self.log_with_context(f"最终标点检查行 {r_fep} 出错: {efp_cell}", r_fep, g_col_idx, "WARNING")
        self.log(f"最终标点检查: {modified_count} 行补充句号" if modified_count else "最终标点检查: 无需补充", "INFO")

    def _copy_intro(self, ws_intro, old_path_intro):
        old_wb_intro_local = None # 使用局部变量
        try:
            self.log(f"尝试从旧表 '{os.path.basename(old_path_intro)}' 复制简介...", "INFO")
            old_wb_intro_local = load_workbook(old_path_intro, data_only=True)
            old_ws_name_intro = '台词管理' if '台词管理' in old_wb_intro_local.sheetnames else old_wb_intro_local.sheetnames[0]
            old_ws_intro_obj = old_wb_intro_local[old_ws_name_intro]
            g_col_idx_old_intro = self._col2idx(self.config['col_g'])
            intro_text_val = old_ws_intro_obj.cell(row=2, column=g_col_idx_old_intro).value or ''
            intro_text_str_val = str(intro_text_val).strip()
            if not intro_text_str_val:
                self.log(f"旧表 '{old_ws_name_intro}' 第2行G列无简介。", "WARNING")
                old_wb_intro_local.close()
                return
            
            # 找出实际的简介内容（最后一个译名后的内容）
            sentences = intro_text_str_val.split("。")
            
            # 查找最后一个"译名是"的位置
            last_translation_index = -1
            for i, sentence in enumerate(sentences):
                if "译名是" in sentence:
                    last_translation_index = i
            
            # 提取最后一个译名之后的内容作为实际简介
            actual_intro = ""
            if last_translation_index != -1 and last_translation_index < len(sentences) - 1:
                actual_intro = "。".join(sentences[last_translation_index + 1:]).strip()
                if actual_intro and not actual_intro.endswith("。"):
                    actual_intro += "。"
                self.log(f"找到实际简介: '{actual_intro}'", "DEBUG")
            else:
                # 如果找不到明确的分隔点，使用原逻辑提取最后一句
                sentences_punc = re.findall(r'[^。？！]+[。？！]*', intro_text_str_val)
                actual_intro = sentences_punc[-1].strip() if sentences_punc else intro_text_str_val
                self.log(f"未找到明确分隔，使用最后一句作为简介: '{actual_intro}'", "DEBUG")
            
            # 确保有内容
            if not actual_intro:
                actual_intro = intro_text_str_val
            
            # 写入H列完整内容
            intro_dest_h_idx = self._col2idx('H')
            ws_intro.Cells(2, intro_dest_h_idx).Value = intro_text_str_val
            self.log_with_context(f"简介完整内容复制到H列:'{intro_text_str_val[:50]}'", 2, intro_dest_h_idx, "DEBUG")
            
            # 将提取出的实际简介写入其他列
            for col_letter in ['J', 'L', 'N', 'P', 'R', 'T', 'V', 'X']:
                col_idx = self._col2idx(col_letter)
                ws_intro.Cells(2, col_idx).Value = actual_intro
            
            old_wb_intro_local.close()
            self.log(f"简介从 '{os.path.basename(old_path_intro)}' 复制完成。实际简介: '{actual_intro[:50]}'", "INFO")
        except Exception as e_intro:
            self.log(f"复制简介失败 ('{os.path.basename(old_path_intro)}'): {e_intro}", "ERROR")
            self.log(traceback.format_exc(), "DEBUG")
        finally:
            if old_wb_intro_local and hasattr(old_wb_intro_local, 'close') and not getattr(old_wb_intro_local, 'closed', True):
                try: old_wb_intro_local.close()
                except Exception as e_close_intro_wb: self.log(f"关闭简介用旧工作簿时出错: {e_close_intro_wb}", "ERROR")


    # ==============================================================================
    # BEGIN: _copy_speakers and its helper methods (FULL IMPLEMENTATION)
    # ==============================================================================
    def _build_character_patterns(self, old_data_list_build):
        speakers_dialogs = {}
        for item in old_data_list_build:
            speaker = item.get('speaker')
            dialog_text = item.get('dialog', '')
            if not speaker or not dialog_text: continue
            if speaker not in speakers_dialogs: speakers_dialogs[speaker] = []
            speakers_dialogs[speaker].append(dialog_text)

        character_patterns = {}
        for speaker, dialogs in speakers_dialogs.items():
            phrases = []
            phrase_counts = {}
            for dialog_text_item in dialogs:
                if len(dialog_text_item) < 2: continue
                for i_pattern in range(len(dialog_text_item) - 1): 
                    for j_pattern_len in range(2, min(6, len(dialog_text_item) - i_pattern + 1)):
                        phrase = dialog_text_item[i_pattern : i_pattern+j_pattern_len]
                        if re.search(r'[\u4e00-\u9fff]', phrase): 
                            phrases.append(phrase)
            
            if not phrases: continue
            for phrase in phrases: phrase_counts[phrase] = phrase_counts.get(phrase, 0) + 1
            
            sorted_phrases = sorted(phrase_counts.items(), key=lambda x: x[1] * len(x[0]), reverse=True)
            top_n_phrases = [p[0] for p in sorted_phrases[:15] if p[1] > 1 and len(p[0]) > 1] 
            if top_n_phrases:
                character_patterns[speaker] = top_n_phrases
                self.log(f"为说话人 '{speaker}' 构建特征模式: {top_n_phrases}", "TRACE")
        return character_patterns

    def _guess_speaker_from_content(self, dialog_text_guess, char_patterns_guess):
        if not dialog_text_guess or not char_patterns_guess: return None, 0.0
        best_match_speaker, highest_score = None, 0.0
        min_score_threshold = 0.3 

        for speaker, patterns in char_patterns_guess.items():
            current_score = 0.0
            for pattern in patterns:
                if pattern in dialog_text_guess:
                    current_score += len(pattern) / 10.0 
            if current_score > highest_score and current_score >= min_score_threshold:
                highest_score = current_score
                best_match_speaker = speaker
        
        final_confidence_threshold = self.config.get('content_guess_confidence_threshold', 0.7)
        return (best_match_speaker, highest_score) if highest_score >= final_confidence_threshold else (None, 0.0)

    def _has_ending_punctuation(self, text):
        """判断文本是否以标点符号结尾"""
        if not text or not isinstance(text, str): return False
        valid_trailing_punctuations = ('。', '！', '？', '.', '!', '?', '；', ';', '：', ':', ')', '）', '"', "'", '"', "'", ']', '】', '》', '…')
        return any(text.endswith(p) for p in valid_trailing_punctuations)

    def _remove_ending_punctuation(self, text):
        """移除文本末尾的标点符号"""
        if not text or not isinstance(text, str): return text
        valid_trailing_punctuations = ('。', '！', '？', '.', '!', '?', '；', ';', '：', ':', ')', '）', '"', "'", '"', "'", ']', '】', '》', '…')
        result = text
        while any(result.endswith(p) for p in valid_trailing_punctuations) and result:
            for punct in valid_trailing_punctuations:
                if result.endswith(punct):
                    result = result[:-len(punct)].rstrip()
                    break
        return result

    # _has_ending_punctuation and _remove_ending_punctuation are already defined globally in the class

    def _process_paragraph_punctuation(self, ws_proc_para, new_data_list_para, dialog_col_idx_lqa):
        self.log("说话人复制后：开始段落标点最终处理...", "INFO")
        paragraphs_map = {} 
        for item_p in new_data_list_para:
            para_id_p = item_p.get('paragraph_id')
            speaker_p = item_p.get('speaker_after_match') 
            if not para_id_p or not speaker_p: continue

            para_key = (speaker_p, para_id_p) 
            if para_key not in paragraphs_map: paragraphs_map[para_key] = []
            paragraphs_map[para_key].append(item_p)
        
        if not paragraphs_map: self.log("段落标点后处理：无有效段落数据。", "INFO"); return

        processed_punct_count = 0
        for para_key, items_in_para_p in paragraphs_map.items():
            if not items_in_para_p or not self.running : continue
            items_in_para_p.sort(key=lambda x: x.get('paragraph_position', float('inf'))) 

            num_items_p = len(items_in_para_p)
            for i_p_item, item_to_punct in enumerate(items_in_para_p):
                is_last_line_p = (i_p_item == num_items_p - 1)
                row_p_num = item_to_punct.get('row')
                if not row_p_num: continue

                try:
                    cell_obj_p = ws_proc_para.Cells(row_p_num, dialog_col_idx_lqa)
                    cell_text_val_p = str(cell_obj_p.Value or '').strip()
                    if not cell_text_val_p: continue

                    has_punct_p = self._has_ending_punctuation(cell_text_val_p)
                    if is_last_line_p: 
                        if not has_punct_p: 
                            cell_obj_p.Value = cell_text_val_p + "。"
                            processed_punct_count += 1
                            self.log_with_context(f"段后标点: 末行 '{cell_text_val_p[:30]}...' 加句号", row=row_p_num, col_idx=dialog_col_idx_lqa, level="TRACE")
                    else: 
                        if has_punct_p: 
                            stripped_text_p = self._remove_ending_punctuation(cell_text_val_p)
                            if stripped_text_p != cell_text_val_p : 
                                cell_obj_p.Value = stripped_text_p
                                processed_punct_count += 1
                                self.log_with_context(f"段后标点: 中行 '{cell_text_val_p[:30]}...' 移除标点后为 '{stripped_text_p[:30]}'", row=row_p_num, col_idx=dialog_col_idx_lqa, level="TRACE")
                except Exception as e_para_punct_cell:
                    self.log_with_context(f"段落标点后处理单元格操作出错: {e_para_punct_cell}", row=row_p_num, col_idx=dialog_col_idx_lqa, level="WARNING")
        self.log(f"说话人复制后的段落标点最终处理完成，共修改 {processed_punct_count} 处。", "INFO")

    def _calculate_segment_coherence(self, segment_texts_list_coh):
        if not segment_texts_list_coh or len(segment_texts_list_coh) <= 1: return 1.0
        coherence_score_val = 0.0; num_transitions_coh = len(segment_texts_list_coh) - 1
        
        for i_coh in range(num_transitions_coh):
            current_text_coh = segment_texts_list_coh[i_coh]
            next_text_coh = segment_texts_list_coh[i_coh+1]
            transition_score_val = 0.0
            if not current_text_coh or not next_text_coh: continue

            continuity_words = ['然后','接着','所以','但是','而且','因为','不过','如果','因此','还有','另外','同时','于是','那么','此外']
            if any(next_text_coh.startswith(word) for word in continuity_words):
                transition_score_val += 0.4

            try:
                current_words_set_coh = set(re.findall(r'[\u4e00-\u9fffA-Za-z0-9]+', current_text_coh))
                next_words_set_coh = set(re.findall(r'[\u4e00-\u9fffA-ZaZ0-9]+', next_text_coh))
                common_words_coh = current_words_set_coh.intersection(next_words_set_coh)
                if common_words_coh:
                     transition_score_val += min(0.3, len(common_words_coh) * 0.05 + 0.05) 
            except TypeError: pass
            coherence_score_val += transition_score_val
        
        max_possible_score_per_trans = 0.7 
        max_total_score_coh = num_transitions_coh * max_possible_score_per_trans
        
        final_coherence = 0.0
        if max_total_score_coh > 0:
            final_coherence = min(1.0, coherence_score_val / max_total_score_coh) if coherence_score_val > 0 else 0.0
        self.log(f"段落连贯性计算: 文本='{'|'.join(s[:10] for s in segment_texts_list_coh)}', 得分={final_coherence:.2f}", "TRACE")
        return final_coherence

    def _looks_like_different_speaker(self, text1_lds, text2_lds):
        if not text1_lds or not text2_lds: return False
        if (text1_lds.endswith('?') or text1_lds.endswith('？')) and \
           not any(text2_lds.startswith(pronoun) for pronoun in ['你','我','他','她']) and len(text2_lds)>2 :
            return True
        pronouns1 = {w for w in ['我','你','您'] if w in text1_lds}
        pronouns2 = {w for w in ['我','你','您'] if w in text2_lds}
        if (('我' in pronouns1 and ('你' in pronouns2 or '您' in pronouns2)) or \
            (('你' in pronouns1 or '您' in pronouns1) and '我' in pronouns2)) and \
           not pronouns1.intersection(pronouns2):
            return True
        if (text1_lds.endswith('吧') or text1_lds.endswith('啊') or text1_lds.endswith('呢')) and \
           not (text2_lds.endswith('吧') or text2_lds.endswith('啊') or text2_lds.endswith('呢')):
            if len(text1_lds) < 10 and len(text2_lds) > 3 : 
                 return True
        return False

    def _looks_like_Youtube(self, row1_lqa, row2_lqa, ws_lqa, dialog_col_idx_lqa):
        try:
            text1_lqa_val = str(ws_lqa.Cells(row1_lqa, dialog_col_idx_lqa).Value or '').strip()
            text2_lqa_val = str(ws_lqa.Cells(row2_lqa, dialog_col_idx_lqa).Value or '').strip()
            if not text1_lqa_val or not text2_lqa_val: return False
            is_q1 = text1_lqa_val.endswith(('?', '？')) or \
                      any(text1_lqa_val.startswith(qw) for qw in ["什么","谁","哪","怎么","为啥","几时","难道","可否","能否"])
            is_a2_start = any(text2_lqa_val.startswith(aw) for aw in ["是","不是","对","没错","好","嗯","不","没","也许","当然","因为","在于","就是"])
            is_a2 = is_a2_start or (is_q1 and len(text2_lqa_val) > 0 and not (text2_lqa_val.endswith(('?', '？')))) 
            return is_q1 and is_a2
        except Exception as e_lqa_cell:
            self.log_with_context(f"判断问答模式时单元格读取出错: {e_lqa_cell}", row=row1_lqa, level="WARNING")
            return False

    def _handle_special_cases(self, ws_hsc, new_data_list_hsc, speaker_col_f_idx_hsc, dialog_col_g_idx_hsc):
        self.log("开始应用特殊情况处理规则 (如说话人交替修正)...", "INFO")
        modified_by_rules_count = 0
        dialog_pattern_hsc = [] 
        all_speakers_set = set()
        for item_h in new_data_list_hsc:
            if item_h.get('matched') and item_h.get('speaker_after_match') and not item_h.get('content_speaker_guess'):
                try:
                    row_h_val = item_h['row']
                    speaker_h_val = item_h['speaker_after_match']
                    dialog_h_val = str(ws_hsc.Cells(row_h_val, dialog_col_g_idx_hsc).Value or '').strip()
                    if dialog_h_val : 
                        dialog_pattern_hsc.append({'row': row_h_val, 'speaker': speaker_h_val, 'dialog': dialog_h_val})
                        all_speakers_set.add(speaker_h_val)
                except Exception as e_hsc_get:
                     self.log_with_context(f"特殊规则：读取数据时出错: {e_hsc_get}", row=item_h.get('row'), level="WARNING")
        
        if not dialog_pattern_hsc or len(dialog_pattern_hsc) < 2 :
            self.log("特殊规则：无足够数据进行交替模式分析。", "INFO"); return
        
        dialog_pattern_hsc.sort(key=lambda x: x['row']) 

        if 2 <= len(all_speakers_set) <= 3: 
            self.log(f"特殊规则：检测到 {len(all_speakers_set)} 个主要说话人 ({', '.join(list(all_speakers_set))})，尝试交替修正。", "DEBUG")
            speakers_list_hsc = list(all_speakers_set)

            for i_hsc in range(1, len(dialog_pattern_hsc)):
                prev_info_h = dialog_pattern_hsc[i_hsc-1]
                curr_info_h = dialog_pattern_hsc[i_hsc]

                if prev_info_h['speaker'] == curr_info_h['speaker']: 
                    should_switch = self._looks_like_different_speaker(prev_info_h['dialog'], curr_info_h['dialog']) or \
                                    self._looks_like_Youtube(prev_info_h['row'], curr_info_h['row'], ws_hsc, dialog_col_g_idx_hsc)
                    
                    if should_switch:
                        other_speaker_h = None
                        if len(speakers_list_hsc) == 2:
                            other_speaker_h = speakers_list_hsc[0] if curr_info_h['speaker'] == speakers_list_hsc[1] else speakers_list_hsc[1]
                        elif len(speakers_list_hsc) == 3: 
                            possible_others = [s for s in speakers_list_hsc if s != curr_info_h['speaker'] and s != prev_info_h['speaker']]
                            if possible_others: other_speaker_h = possible_others[0]
                            # Fallback: try to get speaker from row before prev_info_h if possible (DATA_START_ROW check needed)
                            elif prev_info_h['row'] > self.config.get('_DATA_START_ROW_CACHE', 1) : # Need a way to get DATA_START_ROW here
                                try:
                                    grand_prev_speaker = str(ws_hsc.Cells(prev_info_h['row']-1, speaker_col_f_idx_hsc).Value or '').strip()
                                    if grand_prev_speaker in speakers_list_hsc and grand_prev_speaker != curr_info_h['speaker']:
                                        other_speaker_h = grand_prev_speaker
                                except: pass 

                        if other_speaker_h and other_speaker_h != curr_info_h['speaker']:
                            try:
                                self.log_with_context(f"特殊规则修正(交替): 从 '{curr_info_h['speaker']}' 改为 '{other_speaker_h}'. 原对话: '{curr_info_h['dialog'][:20]}'", row=curr_info_h['row'], col_idx=speaker_col_f_idx_hsc, level="INFO")
                                ws_hsc.Cells(curr_info_h['row'], speaker_col_f_idx_hsc).Value = other_speaker_h
                                dialog_pattern_hsc[i_hsc]['speaker'] = other_speaker_h 
                                # 修复：使用方法参数new_data_list_hsc，而不是未定义的全局变量new_data_list
                                for item_to_update in new_data_list_hsc: 
                                    if item_to_update.get('row') == curr_info_h['row']:
                                        item_to_update['speaker_after_match'] = other_speaker_h; break
                                modified_by_rules_count += 1
                            except Exception as e_hsc_set:
                                self.log_with_context(f"特殊规则修正时写入Excel出错: {e_hsc_set}", row=curr_info_h['row'], level="WARNING")
        
        if modified_by_rules_count > 0: self.log(f"特殊情况处理规则共修正了 {modified_by_rules_count} 处说话人。", "INFO")
        else: self.log("特殊情况处理：未触发明确的修正。", "INFO")

    def _copy_speakers(self, ws_copy, old_path_copy, data_start_row):
        old_wb_copy = None 
        try:
            self.log(f"开始从旧表 '{os.path.basename(old_path_copy)}' 复制说话人 (数据从第 {data_start_row} 行开始)...", "INFO")
            old_wb_copy = load_workbook(old_path_copy, data_only=True)
            old_sheet_name_copy = '台词管理' if '台词管理' in old_wb_copy.sheetnames else old_wb_copy.sheetnames[0]
            old_sheet_obj = old_wb_copy[old_sheet_name_copy]

            speaker_col_f_idx = self._col2idx(self.config['col_f_speaker'])
            dialog_col_g_idx = self._col2idx(self.config['col_g'])

            old_data_list = []
            self.log(f"旧表 '{old_sheet_name_copy}' 最大行: {old_sheet_obj.max_row}", "DEBUG")
            for r_old_val in range(data_start_row, old_sheet_obj.max_row + 1):
                speaker_val = str(old_sheet_obj.cell(row=r_old_val, column=speaker_col_f_idx).value or '').strip()
                dialog_val = str(old_sheet_obj.cell(row=r_old_val, column=dialog_col_g_idx).value or '').strip()
                if dialog_val: 
                    old_data_list.append({'row': r_old_val, 'speaker': speaker_val, 'dialog': dialog_val, 'used': False})
            
            if not old_data_list:
                self.log(f"旧表 '{old_sheet_name_copy}' 未找到可用台词数据 (检查列F/G和起始行{data_start_row})。", "WARNING")
                return
            self.log(f"旧表数据加载完成，共 {len(old_data_list)} 条有效对话。", "DEBUG")

            char_patterns = self._build_character_patterns(old_data_list)
            self.config['_DATA_START_ROW_CACHE'] = data_start_row # 缓存DATA_START_ROW给_handle_special_cases用

            last_row_new = 1
            try: last_row_new = ws_copy.Cells.Find("*", SearchOrder=win32.constants.xlByRows, SearchDirection=win32.constants.xlPrevious).Row
            except: self.log_with_context("复制说话人：无法确定新表最后行，默认1", level="WARNING")
            
            new_data_list = []
            for r_new_val in range(data_start_row, last_row_new + 1):
                dialog_new_val = str(ws_copy.Cells(r_new_val, dialog_col_g_idx).Value or '').strip()
                if dialog_new_val:
                    item = {'row': r_new_val, 'dialog': dialog_new_val, 'matched': False, 
                            'speaker_after_match': None, 'paragraph_id': None, 'paragraph_position': -1}
                    if char_patterns:
                        guess_s, guess_c = self._guess_speaker_from_content(dialog_new_val, char_patterns)
                        if guess_s : 
                            item['content_speaker_guess'] = guess_s
                            item['content_confidence'] = guess_c
                    new_data_list.append(item)

            if not new_data_list:
                self.log("新表中未找到可用台词数据进行匹配。", "WARNING")
                return
            self.log(f"新表数据加载完成，共 {len(new_data_list)} 条有效对话。", "DEBUG")

            exact_thresh = self.config.get('exact_match_threshold', 0.95)
            len_ratio_thresh = self.config.get('exact_match_length_ratio_threshold', 0.7)
            general_match_thresh = self.config.get('speaker_match_threshold', 0.6)
            seg_sim_w = self.config.get('segment_similarity_weight', 0.6)
            seg_coh_w = self.config.get('segment_coherence_weight', 0.25)
            seg_len_w = self.config.get('segment_length_ratio_weight', 0.15)
            matched_s1, matched_s2_lines, matched_s3 = 0, 0, 0

            self.log("说话人复制 - Stage 1: 精确匹配开始...", "INFO")
            for new_item_s1 in new_data_list:
                if not self.running or new_item_s1['matched'] or not new_item_s1['dialog']: continue
                best_old_match_idx_s1, highest_ratio_s1 = -1, 0.0
                for old_idx_s1, old_item_s1 in enumerate(old_data_list):
                    if old_item_s1['used'] or not old_item_s1['dialog']: continue
                    current_ratio_s1 = difflib.SequenceMatcher(None, new_item_s1['dialog'], old_item_s1['dialog']).ratio()
                    if current_ratio_s1 > highest_ratio_s1: highest_ratio_s1, best_old_match_idx_s1 = current_ratio_s1, old_idx_s1
                if highest_ratio_s1 >= exact_thresh and best_old_match_idx_s1 != -1:
                    old_match_s1 = old_data_list[best_old_match_idx_s1]
                    len_n, len_o = len(new_item_s1['dialog'].replace(" ","")), len(old_match_s1['dialog'].replace(" ",""))
                    if len_n > 0 and len_o > 0 and (min(len_n, len_o) / max(len_n, len_o)) > len_ratio_thresh:
                        try:
                            ws_copy.Cells(new_item_s1['row'], speaker_col_f_idx).Value = old_match_s1['speaker']
                            new_item_s1['speaker_after_match'] = old_match_s1['speaker']
                            new_item_s1['matched'] = True; old_match_s1['used'] = True; matched_s1 += 1
                            self.log_with_context(f"S1 精确匹配: OldR {old_match_s1['row']} (S:{old_match_s1['speaker']}) -> NewR {new_item_s1['row']} (R:{highest_ratio_s1:.2f})", row=new_item_s1['row'], level="DEBUG")
                        except Exception as e_s1w : self.log_with_context(f"S1写入失败:{e_s1w}",row=new_item_s1['row'],level="WARNING")
            self.log(f"说话人复制 - Stage 1: 精确匹配结束. 匹配 {matched_s1} 行.", "INFO")
            if not self.running: self.log("中止于S1后"); return

            # 完全重写的Stage 2分段匹配部分:
            self.log("说话人复制 - Stage 2: 分段匹配开始 (增强版)...", "INFO")
            import time
            start_time_s2 = time.time()
            matched_s2_lines = 0
            max_processing_time_s2 = 60
            processed_count = 0
            
            # 新增：按时间顺序跟踪旧表和新表的位置
            old_position = 0
            new_position = 0
            
            # 构建一对多映射检测
            one_to_many_map = {}
            for old_idx, old_item in enumerate(old_data_list):
                if old_item['used']: continue
                old_text = old_item['dialog']
                if not old_text: continue
                
                # 检测是否可能是多行合并的台词
                sentences = re.split(r'[。！？.!?]', old_text)
                sentences = [s for s in sentences if s.strip()]
                
                if len(sentences) > 1:
                    # 可能是多行合并，记录以备后用
                    one_to_many_map[old_idx] = {
                        'sentences': sentences,
                        'full_text': old_text,
                        'speaker': old_item['speaker'],
                        'matched_new_rows': []
                    }
        
            # 收集连续未匹配段落，考虑时间顺序
            unmatched_segments = []
            segment = []
            
            for i, new_item in enumerate(new_data_list):
                if not new_item['matched']:
                    segment.append(new_item)
                else:
                    if segment:
                        unmatched_segments.append(segment)
                        segment = []
            
            if segment:  # 添加最后一个段落
                unmatched_segments.append(segment)
            
            # 处理一对多映射：寻找旧表一行匹配新表多行
            for old_idx, mapping in one_to_many_map.items():
                if not self.running: break
                
                for segment in unmatched_segments:
                    if len(segment) < 2: continue  # 太短的段落跳过
                    
                    # 检查整个段落是否匹配合并的台词
                    combined_text = " ".join([item['dialog'] for item in segment])
                    similarity = difflib.SequenceMatcher(None, combined_text, mapping['full_text']).ratio()
                    
                    if similarity >= general_match_thresh:
                        # 找到匹配，将相同说话人应用到所有行
                        for idx, item in enumerate(segment):
                            try:
                                row = item['row']
                                ws_copy.Cells(row, speaker_col_f_idx).Value = mapping['speaker']
                                item['speaker_after_match'] = mapping['speaker']
                                item['matched'] = True
                                item['paragraph_id'] = processed_count
                                item['paragraph_position'] = idx
                                matched_s2_lines += 1
                                
                                self.log_with_context(
                                    f"S2 一对多匹配: 旧行合并文本 '{mapping['full_text'][:20]}...' -> " +
                                    f"新行 {row} (相似度:{similarity:.2f})",
                                    row=row, level="DEBUG"
                                )
                            except Exception as e:
                                self.log_with_context(f"S2一对多写入失败:{e}", row=item['row'], level="WARNING")
                        
                        old_data_list[old_idx]['used'] = True
                        processed_count += 1
                        break  # 找到匹配后处理下一个合并台词
            
            # 原有的段落匹配逻辑保留，但增加时间顺序约束
            for segment in unmatched_segments:
                if len(segment) < 2: continue  # 跳过单行段落
                
                # 限制搜索范围：
                search_start = max(0, old_position - 5)  # 从当前位置往前最多5行
                search_end = min(len(old_data_list), old_position + 20)  # 往后最多20行
                
                best_match_indices = []
                best_match_score = 0
                best_match_start = -1
                
                # 在有限范围内搜索最佳匹配
                for start_idx in range(search_start, search_end - len(segment) + 1):
                    if not self.running: break
                    
                    # 检查连续段落是否都未被使用
                    all_available = True
                    for i in range(len(segment)):
                        if start_idx + i >= len(old_data_list) or old_data_list[start_idx + i]['used']:
                            all_available = False
                            break
                    
                    if not all_available:
                        continue
                    
                    # 计算段落相似度
                    old_segment_texts = [old_data_list[start_idx + i]['dialog'] for i in range(len(segment))]
                    
                    # 计算整体段落相似度
                    segment_similarity = 0
                    for i in range(len(segment)):
                        sim = difflib.SequenceMatcher(None, segment[i]['dialog'], old_segment_texts[i]).ratio()
                        segment_similarity += sim
                    segment_similarity /= len(segment)
                    
                    # 计算段落连贯性
                    coherence = self._calculate_segment_coherence([item['dialog'] for item in segment])
                    old_coherence = self._calculate_segment_coherence(old_segment_texts)
                    coherence_diff = abs(coherence - old_coherence)
                    coherence_score = 1.0 - min(coherence_diff, 0.5) * 2  # 归一化为0-1
                    
                    # 计算段落长度比例相似度
                    len_similarity = 1.0
                    for i in range(len(segment)):
                        len_n = len(segment[i]['dialog'])
                        len_o = len(old_segment_texts[i])
                        if len_n > 0 and len_o > 0:
                            len_similarity *= min(len_n, len_o) / max(len_n, len_o)
                    len_similarity = len_similarity ** (1.0 / len(segment))  # 几何平均
                    
                    # 综合评分
                    match_score = (
                        seg_sim_w * segment_similarity +
                        seg_coh_w * coherence_score +
                        seg_len_w * len_similarity
                    )
                    
                    if match_score > best_match_score:
                        best_match_score = match_score
                        best_match_start = start_idx
                        best_match_indices = [start_idx + i for i in range(len(segment))]
                
                # 应用最佳匹配
                if best_match_score >= general_match_thresh and best_match_start != -1:
                    for i, new_item_to_match in enumerate(segment):
                        old_idx = best_match_start + i
                        if old_idx < len(old_data_list):
                            old_item_match = old_data_list[old_idx]
                            
                            try:
                                ws_copy.Cells(new_item_to_match['row'], speaker_col_f_idx).Value = old_item_match['speaker']
                                new_item_to_match['speaker_after_match'] = old_item_match['speaker']
                                new_item_to_match['matched'] = True
                                old_item_match['used'] = True
                                new_item_to_match['paragraph_id'] = processed_count
                                new_item_to_match['paragraph_position'] = i
                                matched_s2_lines += 1
                                
                                self.log_with_context(
                                    f"S2 段落匹配: OldR {old_item_match['row']} (S:{old_item_match['speaker']}) -> " +
                                    f"NewR {new_item_to_match['row']} (整体分数:{best_match_score:.2f})",
                                    row=new_item_to_match['row'], level="DEBUG"
                                )
                            except Exception as e_s2w:
                                self.log_with_context(f"S2写入失败:{e_s2w}", row=new_item_to_match['row'], level="WARNING")
                    
                    processed_count += 1
            
            self.log(f"说话人复制 - Stage 2: 分段匹配结束. 耗时 {time.time() - start_time_s2:.1f}秒, 匹配 {matched_s2_lines} 行.", "INFO")
            if not self.running: self.log("中止于S2后"); return

            self.log("说话人复制 - Stage 3: 有序模糊单行匹配开始...", "INFO")
            unmatched_s3_count = 0
            
            # 重置位置跟踪
            old_position = 0
            for new_item_s3 in new_data_list:
                if not self.running or new_item_s3['matched']: continue
                
                # 在附近范围内寻找最佳匹配
                search_window = 15  # 搜索窗口大小
                best_old_match_idx_s3, highest_ratio_s3 = -1, 0.0
                
                # 优先在时间顺序合理的范围内搜索
                search_start = max(0, old_position - 5)
                search_end = min(len(old_data_list), old_position + search_window)
                
                for old_idx_s3 in range(search_start, search_end):
                    old_item_s3 = old_data_list[old_idx_s3]
                    if old_item_s3['used'] or not old_item_s3['dialog']: continue
                    
                    current_ratio_s3 = difflib.SequenceMatcher(None, new_item_s3['dialog'], old_item_s3['dialog']).ratio()
                    if current_ratio_s3 > highest_ratio_s3:
                        highest_ratio_s3, best_old_match_idx_s3 = current_ratio_s3, old_idx_s3
                
                row_s3 = new_item_s3['row']
                if highest_ratio_s3 >= general_match_thresh and best_old_match_idx_s3 != -1:
                    # 找到有效匹配，更新旧表位置
                    old_match_s3 = old_data_list[best_old_match_idx_s3]
                    old_position = best_old_match_idx_s3 + 1
                    
                    try:
                        ws_copy.Cells(row_s3, speaker_col_f_idx).Value = old_match_s3['speaker']
                        new_item_s3['speaker_after_match'] = old_match_s3['speaker']
                        new_item_s3['matched'] = True
                        old_match_s3['used'] = True
                        matched_s3 += 1
                        self.log_with_context(f"S3 有序匹配: OldR {old_match_s3['row']} -> NewR {row_s3} (R:{highest_ratio_s3:.2f})", row=row_s3, level="DEBUG")
                    except Exception as e_s3w:
                        self.log_with_context(f"S3写入失败:{e_s3w}", row=row_s3, level="WARNING")
                else:
                    # 只对内容猜测更有信心的情况进行标记，降低误标红率
                    unmatched_s3_count += 1
                    try:
                        content_guess_threshold_final = self.config.get('content_guess_confidence_threshold', 0.7)
                        content_guess_threshold_higher = content_guess_threshold_final + 0.1  # 提高阈值
                        if new_item_s3.get('content_speaker_guess') and new_item_s3.get('content_confidence', 0) >= content_guess_threshold_higher:
                            guessed_s_val = new_item_s3['content_speaker_guess']
                            ws_copy.Cells(row_s3, speaker_col_f_idx).Value = guessed_s_val
                            new_item_s3['speaker_after_match'] = guessed_s_val
                            ws_copy.Cells(row_s3, speaker_col_f_idx).Interior.ColorIndex = 6
                            self.log_with_context(f"S3 内容猜测填补: '{guessed_s_val}' (高可信度:{new_item_s3['content_confidence']:.2f})", row=row_s3, level="INFO")
                        else:
                            # 不再标红所有未匹配项，只标记那些特别可疑的
                            suspicious = False
                            for prev_idx in range(max(0, i-1), max(0, i-3), -1):
                                if prev_idx >= 0 and prev_idx < len(new_data_list):
                                    prev_item = new_data_list[prev_idx]
                                    if prev_item.get('matched') and prev_item.get('speaker_after_match'):
                                        # 如果前面几行台词与当前对话模式不符，标记为可疑
                                        if self._looks_like_different_speaker(prev_item['dialog'], new_item_s3['dialog']):
                                            suspicious = True
                                            break
                        
                        if suspicious:
                            ws_copy.Cells(row_s3, speaker_col_f_idx).Interior.ColorIndex = 3
                    except Exception as e_s3color:
                        self.log_with_context(f"S3标色失败:{e_s3color}", row=row_s3, level="WARNING")
            self.log(f"说话人复制 - Stage 3: 模糊单行匹配结束. 匹配 {matched_s3} 行. 未匹配 {unmatched_s3_count} 行.", "INFO")
            if not self.running: 
                self.log("中止于S3后")
                return

            # 这些行必须在try块内，不能在try块之外
            self._process_paragraph_punctuation(ws_copy, new_data_list, dialog_col_g_idx)
            if not self.running: 
                self.log("中止于段落标点后处理后")
                return
                
            self._handle_special_cases(ws_copy, new_data_list, speaker_col_f_idx, dialog_col_g_idx)

            total_matched = matched_s1 + matched_s2_lines + matched_s3
            self.log(f"说话人复制总结: 总匹配行数 {total_matched} (精确:{matched_s1}, 分段行数:{matched_s2_lines}, 模糊:{matched_s3}).", "INFO")
            
        except Exception as e:
            self.log(f"复制说话人主流程发生严重错误: {e}", "CRITICAL")
            self.log(traceback.format_exc(), "DEBUG")
        finally: 
            if old_wb_copy and hasattr(old_wb_copy, 'close') and not getattr(old_wb_copy, 'closed', True):
                try: old_wb_copy.close()
                except Exception as e_close_old_wb: self.log(f"关闭旧工作簿时出错: {e_close_old_wb}", "ERROR")

    def _validate_speaker_assignments(self, ws, new_data_list, speaker_col_idx):
        self.log("进行说话人匹配自检验证...", "INFO")
        issues = 0
        
        # 检查同一角色连续说话行过多
        consecutive_same_speaker = 0
        last_speaker = None
        
        # 按行遍历检查
        for item in sorted(new_data_list, key=lambda x: x['row']):
            current_speaker = item.get('speaker_after_match')
            if not current_speaker:
                continue
                
            # 检查同一角色连续说话
            if current_speaker == last_speaker:
                consecutive_same_speaker += 1
                if consecutive_same_speaker > 7:  # 超过7行同一角色连续说话
                    row = item['row']
                    self.log_with_context(f"自检警告: 角色 '{current_speaker}' 连续说话超过7行", row=row, level="WARNING")
                    # 标记可能的问题
                    try:
                        ws.Cells(row, speaker_col_idx).Interior.ColorIndex = 45  # 使用不同的颜色标记
                        issues += 1
                    except:
                        pass
            else:
                consecutive_same_speaker = 1
                last_speaker = current_speaker
        
        # 检查其他可能的问题...
        
        if issues > 0:
            self.log(f"自检发现 {issues} 处可能的说话人分配问题，已用特别颜色标记", "WARNING")
        else:
            self.log("说话人分配自检通过", "INFO")

    def _col2idx(self, col_letter):
        """将Excel列字母转换为列索引数字"""
        if not col_letter or not isinstance(col_letter, str):
            return 1  # 默认返回第一列
        col_letter = col_letter.upper().strip()
        idx = 0
        for c in col_letter:
            idx = idx * 26 + (ord(c) - ord('A') + 1)
        return idx

    def _idx2col(self, col_idx):
        """将列索引数字转换为Excel列字母"""
        if not isinstance(col_idx, int) or col_idx < 1:
            return 'A'  # 默认返回A列
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def adjust_images(self, ws_images):
        """调整工作表中的图片尺寸"""
        try:
            # 获取工作表中所有图片和形状
            shape_count = ws_images.Shapes.Count
            if shape_count == 0:
                self.log("图片调整: 未发现图片/形状", "INFO")
                return
            
            # 明确设置为1.0，表示保持原尺寸
            scale_factor = 1.0  # 保持原尺寸，不进行缩放
            
            self.log(f"开始处理工作表 '{ws_images.Name}' 中的 {shape_count} 个图片/形状 (保持原尺寸)...", "INFO")
            
            success_count = 0
            failure_count = 0
            
            # 遍历所有形状对象，但不进行缩放处理
            for i in range(1, shape_count + 1):
                if not self.running:
                    break
                    
                try:
                    shape = ws_images.Shapes(i)
                    # 只进行记录，不执行缩放操作
                    original_width = shape.Width
                    original_height = shape.Height
                    self.log(f"图片调整: 形状{i} - 保持原尺寸 {original_width:.1f}x{original_height:.1f}", "TRACE")
                    success_count += 1
                except Exception as e_shape:
                    failure_count += 1
                    self.log(f"图片调整: 形状{i}处理失败: {e_shape}", "WARNING")
                    
            self.log(f"图片调整:成功{success_count},失败{failure_count}", "INFO")
        except Exception as e_all_images:
            self.log(f"处理图片时发生错误: {e_all_images}", "ERROR")

# 主程序入口
if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = ExcelBatchProcessor(root)
        root.mainloop()
    except Exception as e:
        print(f"程序启动失败: {e}")
        import traceback
        traceback.print_exc()
        input("按任意键退出...")  # 在控制台窗口关闭前暂停
